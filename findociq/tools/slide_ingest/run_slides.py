"""run_slides — ingest a CFO-presentation deck: PDF -> Excel + audit + its own DB.

    python3 findociq/tools/slide_ingest/run_slides.py <deck.pdf> --pages 1-30
    python3 findociq/tools/slide_ingest/run_slides.py <deck.pdf> --pages 3 --dry-run

SEPARATE FROM THE STATEMENT PIPELINE, deliberately, on three counts:

  * **No TOC.** `run_doc.py` plans units from a table of contents. A deck has no
    contents page and a slide IS the unit, so pages are the unit boundary and
    STEP 1 has nothing to do. This mirrors the retired `auto_extract.py`, which
    took `--pages 1-30` for the same reason.
  * **Two-pass, image-first.** OBSERVE the rendered page (structure + spatial
    failure modes) -> EXTRACT to HTML guided by that observation. Statement
    extraction is one pass over the PDF bytes; a chart has to be looked at.
  * **Its own database.** Writes `db/compiled_slides.db`, never
    `compiled_fs.db`. Deck data is chart-derived and element-shaped; mixing it
    into `schema_v7` would put figures read off a donut beside figures read off
    a filed statement, in the same `cell_fact`. The two are not the same
    evidence and the schema should not pretend otherwise.

Everything a human needs to audit a run is written per page under `audit/`:
`page.png` (what the model saw), `observe_prompt.txt` / `observe.txt`,
`extract_prompt.txt` / `response.html`, and `parsed.json`.
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import re
import sqlite3
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO / "findociq" / "pipeline"))
sys.path.insert(0, str(HERE))

from html_tables import parse_elements  # noqa: E402

PROMPTS = HERE / "prompts"
DEFAULT_DB = REPO / "findociq" / "db" / "compiled_slides.db"
OUT_ROOT = REPO / "findociq" / "outputs" / "slides"
MODEL = os.environ.get("FINDOCIQ_SLIDES_MODEL", "gemini-2.5-pro")
RENDER_SCALE = 2.5          # matches the retired auto_extract.py

SCHEMA = """
CREATE TABLE IF NOT EXISTS slide_doc (
    doc_id      TEXT PRIMARY KEY,
    source_file TEXT NOT NULL,
    n_pages     INTEGER,
    ingested_at TEXT
);
CREATE TABLE IF NOT EXISTS slide_element (
    doc_id        TEXT NOT NULL,
    page          INTEGER NOT NULL,
    element_idx   INTEGER NOT NULL,
    element_type  TEXT,
    element_title TEXT,
    PRIMARY KEY (doc_id, page, element_idx)
);
CREATE TABLE IF NOT EXISTS slide_cell (
    doc_id      TEXT NOT NULL,
    page        INTEGER NOT NULL,
    element_idx INTEGER NOT NULL,
    row_idx     INTEGER NOT NULL,
    row_label   TEXT,
    col_label   TEXT,
    value_raw   TEXT,
    data_kind   TEXT,
    data_sign   TEXT
);
CREATE INDEX IF NOT EXISTS ix_slide_cell_doc ON slide_cell(doc_id, page);
"""


# ---------------------------------------------------------------------------
# pure helpers
# ---------------------------------------------------------------------------
def parse_pages(spec: str) -> list[int]:
    """'3' / '1-30' / '2,5,9-11' -> sorted unique 1-based page numbers."""
    out: set[int] = set()
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            out.update(range(int(a), int(b) + 1))
        else:
            out.add(int(part))
    return sorted(out)


def build_prompts(observation: str) -> tuple[str, str]:
    """(observe_prompt, extract_prompt). The extract prompt injects the
    observation and appends the shared table rules, exactly as the retired
    auto_extract.py did."""
    observe = (PROMPTS / "slide_observe.txt").read_text(encoding="utf-8")
    extract = (PROMPTS / "slide_extract_html.txt").read_text(encoding="utf-8")
    core = (PROMPTS / "stage2_core.txt").read_text(encoding="utf-8")
    return observe, extract.replace("{OBSERVE_OUTPUT}", observation) + "\n\n" + core


def doc_id_for(pdf: Path) -> str:
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", pdf.stem)


# ---------------------------------------------------------------------------
# IO
# ---------------------------------------------------------------------------
def render_page(pdf: Path, page_1based: int, scale: float = RENDER_SCALE) -> bytes:
    import io
    import pypdfium2 as pdfium
    doc = pdfium.PdfDocument(str(pdf))
    try:
        pil = doc[page_1based - 1].render(scale=scale).to_pil()
        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        return buf.getvalue()
    finally:
        doc.close()


def call_model(client, png: bytes, prompt: str) -> tuple[str, dict]:
    from google.genai import types
    resp = client.models.generate_content(
        model=MODEL,
        contents=[types.Part.from_bytes(data=png, mime_type="image/png"), prompt],
    )
    u = getattr(resp, "usage_metadata", None)
    usage = {
        "prompt_tokens": getattr(u, "prompt_token_count", 0) or 0,
        "output_tokens": getattr(u, "candidates_token_count", 0) or 0,
    } if u else {}
    return (resp.text or ""), usage


def write_excel(elements_by_page: dict[int, list[dict]], out_xlsx: Path) -> None:
    """One sheet per (page, element). Sheet name is p<N>_<idx>_<type> truncated
    to Excel's 31-char limit; the full title is written in the first row so it
    is never lost to truncation."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for page in sorted(elements_by_page):
        for idx, el in enumerate(elements_by_page[page], start=1):
            name = f"p{page}_{idx}_{el['element_type']}"[:31]
            ws = wb.create_sheet(title=name)
            ws.append([el.get("element_title") or ""])
            ws.append([])
            ws.append(["", *el.get("columns", [])])
            for r in el.get("rows", []):
                marks = [m for m in (r.get("kind"), r.get("sign")) if m]
                label = f"{r['label']}  [{' '.join(marks)}]" if marks else r["label"]
                ws.append([label, *r.get("cells", [])])
    if not wb.sheetnames:
        wb.create_sheet(title="empty")
    wb.save(out_xlsx)


def load_db(db: Path, doc_id: str, source_file: str,
            elements_by_page: dict[int, list[dict]]) -> tuple[int, int]:
    """Replace this doc in `db`. Doc-scoped, like the statement loader — a
    re-run cannot double a deck."""
    con = sqlite3.connect(db)
    try:
        con.executescript(SCHEMA)
        cur = con.cursor()
        for t in ("slide_cell", "slide_element", "slide_doc"):
            cur.execute(f"DELETE FROM {t} WHERE doc_id = ?", (doc_id,))
        cur.execute(
            "INSERT INTO slide_doc(doc_id, source_file, n_pages, ingested_at) "
            "VALUES (?,?,?,?)",
            (doc_id, source_file, len(elements_by_page),
             _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")))
        n_el = n_cell = 0
        for page in sorted(elements_by_page):
            for idx, el in enumerate(elements_by_page[page], start=1):
                cur.execute(
                    "INSERT INTO slide_element(doc_id,page,element_idx,"
                    "element_type,element_title) VALUES (?,?,?,?,?)",
                    (doc_id, page, idx, el["element_type"], el["element_title"]))
                n_el += 1
                cols = el.get("columns") or []
                for r_i, r in enumerate(el.get("rows", []), start=1):
                    cells = r.get("cells") or []
                    for c_i, val in enumerate(cells):
                        col = cols[c_i] if c_i < len(cols) else f"col{c_i + 1}"
                        cur.execute(
                            "INSERT INTO slide_cell(doc_id,page,element_idx,row_idx,"
                            "row_label,col_label,value_raw,data_kind,data_sign) "
                            "VALUES (?,?,?,?,?,?,?,?,?)",
                            (doc_id, page, idx, r_i, r["label"], col, val,
                             r.get("kind"), r.get("sign")))
                        n_cell += 1
        con.commit()
        return n_el, n_cell
    finally:
        con.close()


# ---------------------------------------------------------------------------
def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("pdf")
    ap.add_argument("--pages", default="1", help="'3' | '1-30' | '2,5,9-11'")
    ap.add_argument("--tag", default=None,
                    help="output dir under outputs/slides/ (default: the PDF stem)")
    ap.add_argument("--db", default=str(DEFAULT_DB),
                    help=f"slides DB, SEPARATE from compiled_fs.db (default {DEFAULT_DB.name})")
    ap.add_argument("--no-db", action="store_true", help="write Excel + audit only")
    ap.add_argument("--dry-run", action="store_true",
                    help="render pages and assemble prompts, make NO API call")
    args = ap.parse_args(argv)

    pdf = Path(args.pdf)
    if not pdf.exists():
        pdf_alt = REPO / args.pdf
        if not pdf_alt.exists():
            print(f"PDF not found: {args.pdf}", file=sys.stderr)
            return 2
        pdf = pdf_alt

    pages = parse_pages(args.pages)
    doc_id = doc_id_for(pdf)
    run_dir = OUT_ROOT / (args.tag or doc_id)
    audit_root = run_dir / "audit" / doc_id
    (run_dir / "logs").mkdir(parents=True, exist_ok=True)
    audit_root.mkdir(parents=True, exist_ok=True)

    print(f"deck   : {pdf}")
    print(f"pages  : {pages}")
    print(f"out    : {run_dir}")
    print(f"db     : {'(none)' if args.no_db else args.db}")

    client = None
    if not args.dry_run:
        from gemini_client import build_client
        client = build_client()

    elements_by_page: dict[int, list[dict]] = {}
    usage_total: dict[str, int] = {}

    for page in pages:
        pdir = audit_root / f"p{page}"
        pdir.mkdir(parents=True, exist_ok=True)
        png = render_page(pdf, page)
        (pdir / "page.png").write_bytes(png)

        observe_prompt, _ = build_prompts("")
        (pdir / "observe_prompt.txt").write_text(observe_prompt, encoding="utf-8")

        if args.dry_run:
            _, extract_preview = build_prompts("<OBSERVATION GOES HERE>")
            (pdir / "extract_prompt.txt").write_text(extract_preview, encoding="utf-8")
            print(f"  [dry-run] p{page}: rendered {len(png):,}B, prompts written, no API call")
            continue

        observation, u1 = call_model(client, png, observe_prompt)
        (pdir / "observe.txt").write_text(observation, encoding="utf-8")

        _, extract_prompt = build_prompts(observation)
        (pdir / "extract_prompt.txt").write_text(extract_prompt, encoding="utf-8")
        html, u2 = call_model(client, png, extract_prompt)
        (pdir / "response.html").write_text(html, encoding="utf-8")

        for k in set(u1) | set(u2):
            usage_total[k] = usage_total.get(k, 0) + u1.get(k, 0) + u2.get(k, 0)

        els = parse_elements(html)
        elements_by_page[page] = els
        (pdir / "parsed.json").write_text(
            json.dumps({"page": page, "elements": els}, indent=2), encoding="utf-8")
        kinds = ", ".join(e["element_type"] for e in els) or "none"
        n_cells = sum(len(r.get("cells") or []) for e in els for r in e.get("rows", []))
        print(f"  p{page}: {len(els)} element(s) [{kinds}], {n_cells} cell(s)")

    if args.dry_run:
        print("\n[dry-run] nothing called, nothing written beyond prompts + page renders.")
        return 0

    xlsx = run_dir / f"{doc_id}_slides.xlsx"
    write_excel(elements_by_page, xlsx)
    index = {
        "doc_id": doc_id, "source_file": str(pdf), "pages": pages,
        "elements": {str(p): [e["element_type"] for e in els]
                     for p, els in elements_by_page.items()},
        "usage": usage_total,
    }
    (run_dir / f"{doc_id}_slides.index.json").write_text(
        json.dumps(index, indent=2), encoding="utf-8")
    (run_dir / "logs" / "cost_summary.json").write_text(
        json.dumps({"model": MODEL, "usage": usage_total}, indent=2), encoding="utf-8")

    print(f"\n  Excel : {xlsx}")
    print(f"  audit : {audit_root}")
    if not args.no_db:
        n_el, n_cell = load_db(Path(args.db), doc_id, str(pdf), elements_by_page)
        print(f"  DB    : {args.db}  ({n_el} elements, {n_cell} cells)")
    print(f"  tokens: {usage_total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
