"""tag_workbook.py — Flow 1 tail: DB rows -> per-table-family concept-tagging Excel.

Reads what extraction already loaded into `compiled_fs.db` (row_dim / cell_fact /
col_dim) and emits a workbook finance uses to apply concept identities. The
machine pre-fills everything derivable (label, value, unit, AI-proposed concept,
agg_role, group); finance later fills concept identity by hand and re-uploads
(Flow 2, `tag_ingest.py`, not built yet).

Design: findociq/docs/2026-07-29-tag-workbook-design.md
Reference format: ~/Downloads/FinDocIQ_Tag_Highlights_FY25.xlsx (finance's
hand-built example — this module reproduces its layout exactly for the
income-statement / FY2025 slice).

This module does NOT reimplement period or accounting-nature logic — it only
reads what the existing pipeline (run_doc STEP 0-3, concept resolution) has
already written to the DB.
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

PIPELINE_DIR = Path(__file__).resolve().parent
CONCEPT_DICT_PATH = PIPELINE_DIR / "concept" / "concept_dictionary.yaml"

# ---------------------------------------------------------------------------
# Registry (hardcoded for this first slice — income statement / FY2025 / 3 banks)
# ---------------------------------------------------------------------------

FAMILY = "income_statement"

SLICE = {  # family -> per-bank (doc_id, table_id)
    "DBS": (
        "DBS_4Q25_performance_summary",
        "overview_selected_income_statement_items_m_2025-12-31",
    ),
    "OCBC": (
        "OCBC_4Q25_Condensed_Financial_Statements",
        "consolidated_income_statement_consolidated_income_statement_for_the_"
        "financial_year_ended_31_december_2025_2025-12-31",
    ),
    "UOB": (
        "UOB_4Q25_condensed-financial-statements",
        "income_statement_audited_for_the_financial_year_ended_31_income_"
        "statement_audited_for_the_financial_year_ended_31_december_2025_"
        "2025-12-31",
    ),
}

TAG_SHEET_NAME = "TAG — Income Statement FY25"
TAG_HEADERS = [
    "Bank", "#", "Blk", "Line label as printed", "FY2025", "Unit",
    "Concept ID", "Agg role", "Group #", "Note / question",
]

# Style constants, matched to the reference workbook
NAVY = "FF0A2540"
SLATE = "FF44586B"
GREEN_NEW = "FFE8F3EC"
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BANK_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
BANK_HEADER_FILL = PatternFill("solid", fgColor=SLATE)
TITLE_FONT = Font(name="Arial", size=18, bold=True, color=NAVY)
BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrapText=True)

# Row-filter noise patterns
_FOOTNOTE_PREFIX = "notes:"
_NUMERIC_ONLY_RE = re.compile(r"^[\d,\s()]+$")

# FY-column selection regex/rules
_PCT_CHG_LABEL_RE = re.compile(r"%|chg", re.IGNORECASE)
_EXCLUDED_SPANS = {"1H", "2H", "Q1", "Q2", "Q3", "Q4"}
_PREFERRED_LABEL_RE = re.compile(
    r"year ?2025|fy ?2025|full year.*2025|financial year.*2025|12 ?months?.*2025",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Pure helpers (unit-tested without DB / file I/O)
# ---------------------------------------------------------------------------

def is_noise_row(label: str | None) -> bool:
    """True if a row_leaf_label is empty, a footnote block header, or a
    mis-extracted numeric-only row (should be skipped from the tag sheet)."""
    if label is None:
        return True
    stripped = label.strip()
    if not stripped:
        return True
    if stripped.lower().startswith(_FOOTNOTE_PREFIX):
        return True
    if _NUMERIC_ONLY_RE.match(stripped):
        return True
    return False


def derive_agg_and_groups(rows: list[tuple[int, int | None]]) -> dict[int, tuple[str, str | None]]:
    """Given [(row_id, sums_to), ...], return {row_id: (agg_role, group_label)}.

    - `total`     if row_id is itself a sums_to target of some other row
    - `component` elif sums_to is not None (rolls up into a total)
    - `atomic`    otherwise

    Group #: one G-label per distinct sums_to target, assigned in first-
    appearance order (scanning rows in input order). A component row gets its
    target's group; the target (total) row gets that same group. Atomic rows
    get no group (None).
    """
    targets = {sums_to for _row_id, sums_to in rows if sums_to is not None}

    group_of_target: dict[int, str] = {}
    next_group_num = 1
    for _row_id, sums_to in rows:
        if sums_to is not None and sums_to not in group_of_target:
            group_of_target[sums_to] = f"G{next_group_num}"
            next_group_num += 1

    result: dict[int, tuple[str, str | None]] = {}
    for row_id, sums_to in rows:
        if row_id in targets:
            role = "total"
            group = group_of_target.get(row_id)
        elif sums_to is not None:
            role = "component"
            group = group_of_target.get(sums_to)
        else:
            role = "atomic"
            group = None
        result[row_id] = (role, group)
    return result


def select_fy_column(cols: list[dict]) -> tuple[dict | None, str | None]:
    """Pick the full-year-2025 column from a list of col_dim-shaped dicts
    (keys: col_id, col_leaf_label, col_period, period_span, unit).

    Returns (chosen_col_or_None, warning_or_None).

    NOTE: the full-year column's col_period/period_span are often NULL in the
    data today (a period-resolution gap upstream) — that's why label matching
    is the PRIMARY selector here, not col_period/period_span. Known follow-up:
    once period resolution is fixed for these columns, this can lean more on
    col_period/period_span and less on label regex.
    """
    candidates = []
    for col in cols:
        label = (col.get("col_leaf_label") or "")
        unit = col.get("unit")
        span = col.get("period_span")
        if unit == "%":
            continue
        if _PCT_CHG_LABEL_RE.search(label):
            continue
        if span in _EXCLUDED_SPANS:
            continue
        if "2024" in label:
            continue
        candidates.append(col)

    if not candidates:
        return None, "no candidate FY2025 column found (all excluded)"

    preferred = [
        col for col in candidates
        if _PREFERRED_LABEL_RE.search(col.get("col_leaf_label") or "")
        or (col.get("col_period") == "2025-12-31" and col.get("period_span") in ("FY", "12M", "Y"))
    ]
    if len(preferred) == 1:
        return preferred[0], None
    if len(preferred) > 1:
        return preferred[0], (
            f"multiple preferred FY2025 columns matched ({[c.get('col_leaf_label') for c in preferred]}); "
            f"picked first"
        )

    if len(candidates) == 1:
        return candidates[0], None

    return candidates[0], (
        f"ambiguous FY2025 column selection among {[c.get('col_leaf_label') for c in candidates]}; picked first"
    )


def format_value(value_num, value_raw):
    """Cell value to write for the FY2025 column: prefer numeric, fall back
    to the raw extracted text."""
    if value_num is not None:
        return value_num
    return value_raw


# ---------------------------------------------------------------------------
# Concept dictionary
# ---------------------------------------------------------------------------

def load_concept_dictionary(path: Path = CONCEPT_DICT_PATH) -> list[tuple[str, str]]:
    """Return [(concept_key, name), ...] in file order."""
    with open(path) as f:
        doc = yaml.safe_load(f)
    out = []
    for entry in doc.get("concepts", []):
        out.append((entry["key"], entry.get("name", "")))
    return out


# ---------------------------------------------------------------------------
# DB reads
# ---------------------------------------------------------------------------

def fetch_bank_data(con: sqlite3.Connection, doc_id: str, table_id: str) -> dict:
    """Read row_dim + col_dim + cell_fact for one (doc_id, table_id) and
    return {"rows": [...], "fy_col": {...}|None, "warning": str|None}."""
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    rows = [
        dict(r) for r in cur.execute(
            "SELECT row_id, row_leaf_label, line_no, concept_key, sums_to, unit "
            "FROM row_dim WHERE doc_id=? AND table_id=? ORDER BY row_id",
            (doc_id, table_id),
        )
    ]
    cols = [
        dict(r) for r in cur.execute(
            "SELECT col_id, col_leaf_label, col_period, period_span, unit "
            "FROM col_dim WHERE doc_id=? AND table_id=? ORDER BY col_id",
            (doc_id, table_id),
        )
    ]

    fy_col, warning = select_fy_column(cols)

    fy_values: dict[int, object] = {}
    fy_unit: dict[int, object] = {}
    if fy_col is not None:
        for r in cur.execute(
            "SELECT row_id, value_num, value_raw, unit FROM cell_fact "
            "WHERE doc_id=? AND table_id=? AND col_id=?",
            (doc_id, table_id, fy_col["col_id"]),
        ):
            fy_values[r["row_id"]] = format_value(r["value_num"], r["value_raw"])
            fy_unit[r["row_id"]] = r["unit"]

    return {
        "rows": rows,
        "fy_values": fy_values,
        "fy_unit": fy_unit,
        "fy_col": fy_col,
        "warning": warning,
    }


# ---------------------------------------------------------------------------
# Workbook emission
# ---------------------------------------------------------------------------

def build_workbook(db_path: Path, family: str = FAMILY) -> tuple[openpyxl.Workbook, dict]:
    if family != FAMILY:
        raise ValueError(f"unsupported family {family!r}; only {FAMILY!r} is wired for this slice")

    con = sqlite3.connect(str(db_path))
    try:
        per_bank = {}
        warnings: list[str] = []
        for bank, (doc_id, table_id) in SLICE.items():
            data = fetch_bank_data(con, doc_id, table_id)
            if not data["rows"]:
                raise RuntimeError(
                    f"BLOCKED: no row_dim rows for {bank} (doc_id={doc_id!r}, table_id={table_id!r})"
                )
            if data["warning"]:
                warnings.append(f"{bank}: {data['warning']}")
            per_bank[bank] = data
    finally:
        con.close()

    concept_dict = load_concept_dictionary()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    counts = _build_start_here(wb)
    tag_counts, last_data_row = _build_tag_sheet(wb, per_bank, concept_dict)
    _build_concept_dictionary(wb, concept_dict)
    _build_coverage_check(wb, last_data_row)

    return wb, {"per_bank": tag_counts, "warnings": warnings}


def _build_start_here(wb: openpyxl.Workbook):
    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 3.0
    ws.column_dimensions["B"].width = 108.0
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Income Statement — FY2025 / 4Q25"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Table family: Income statement. DBS, OCBC and UOB in one pass."

    ws["B5"] = "WHAT THIS IS"
    ws["B5"].font = Font(name="Arial", bold=True)
    ws["B6"] = (
        "Every printed income-statement line from all three banks, in the order it "
        "appears in the source document. Tick down the PDF alongside this sheet — "
        "if a line is missing here, the extraction missed it and we need to know."
    )
    ws["B8"] = (
        "The FY2025 column is shown so you can eyeball each row against the "
        "published page. You are tagging the LINE, not that number — the tag "
        "applies to every period and every future quarter of this table."
    )

    ws["B10"] = "WHAT TO DO"
    ws["B10"].font = Font(name="Arial", bold=True)
    ws["B11"] = (
        "Most rows are pre-filled with a proposed identity from the concept "
        "dictionary. Read down and correct anything wrong. Rows left blank are "
        "ones only you can settle."
    )
    ws["B12"] = "Agg role: total = the line others sum into · component = part of a total · atomic = stands alone."
    ws["B13"] = "Group #: rows that belong to the same total share a group number, across banks."

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None and cell.coordinate != "B2":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    return {}


def _bank_section_header(bank: str) -> str:
    return f"{bank} — Income statement"


def _build_tag_sheet(wb: openpyxl.Workbook, per_bank: dict, concept_dict: list[tuple[str, str]]):
    ws = wb.create_sheet(TAG_SHEET_NAME)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Income Statement FY2025 — comprehensive line tagging"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = "Every printed row, in source order. Correct anything wrong; fill anything blank."

    for col_idx, header in enumerate(TAG_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    widths = [7.0, 5.0, 6.0, 54.0, 12.0, 7.0, 34.0, 12.0, 9.0, 44.0]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    counts: dict[str, dict] = {}
    r = 5
    first_data_row = r
    for bank, data in per_bank.items():
        ws.cell(row=r, column=1, value=_bank_section_header(bank)).font = BANK_HEADER_FONT
        ws.cell(row=r, column=1).fill = BANK_HEADER_FILL
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = BANK_HEADER_FILL
        r += 1

        rows = [row for row in data["rows"] if not is_noise_row(row["row_leaf_label"])]
        agg_groups = derive_agg_and_groups([(row["row_id"], row["sums_to"]) for row in rows])

        tagged = 0
        running = 0
        for row in rows:
            running += 1
            role, group = agg_groups[row["row_id"]]
            value = data["fy_values"].get(row["row_id"])
            unit = data["fy_unit"].get(row["row_id"]) or row["unit"]
            concept_key = row["concept_key"]
            if concept_key:
                tagged += 1

            ws.cell(row=r, column=1, value=bank)
            ws.cell(row=r, column=2, value=running)
            ws.cell(row=r, column=3, value="IS")
            ws.cell(row=r, column=4, value=row["row_leaf_label"])
            ws.cell(row=r, column=5, value=value)
            ws.cell(row=r, column=6, value=unit)
            ws.cell(row=r, column=7, value=concept_key)
            ws.cell(row=r, column=8, value=role)
            ws.cell(row=r, column=9, value=group)
            ws.cell(row=r, column=10, value=None)
            r += 1

        counts[bank] = {"rows": len(rows), "tagged": tagged}

    last_data_row = r - 1

    concept_dv = DataValidation(
        type="list",
        formula1=f"='Concept Dictionary'!$A$5:$A${4 + len(concept_dict)}",
        allow_blank=True,
    )
    concept_dv.add(f"G{first_data_row}:G{last_data_row}")
    ws.add_data_validation(concept_dv)

    agg_dv = DataValidation(type="list", formula1='"total,component,atomic"', allow_blank=True)
    agg_dv.add(f"H{first_data_row}:H{last_data_row}")
    ws.add_data_validation(agg_dv)

    ws.freeze_panes = "E5"
    return counts, last_data_row


def _build_concept_dictionary(wb: openpyxl.Workbook, concept_dict: list[tuple[str, str]]):
    ws = wb.create_sheet("Concept Dictionary")
    ws.column_dimensions["A"].width = 38.0
    ws.column_dimensions["B"].width = 58.0
    ws.column_dimensions["C"].width = 8.0
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Concept dictionary"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = "Green = new, introduced by this table family. Please sanity-check the names."

    headers = ["Concept ID", "Meaning", "New?"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, (key, name) in enumerate(concept_dict):
        r = 5 + i
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=name)
        # "New?" left blank for now — no baseline to diff against yet.
        ws.cell(row=r, column=3, value=None)
    return {}


def _build_coverage_check(wb: openpyxl.Workbook, last_data_row: int):
    ws = wb.create_sheet("Coverage check")
    ws.column_dimensions["A"].width = 12.0
    ws.column_dimensions["B"].width = 17.0
    ws.column_dimensions["C"].width = 12.0
    ws.column_dimensions["D"].width = 13.0
    ws.column_dimensions["E"].width = 24.0
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Coverage check"
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws["A2"] = "Compare against the printed pages. If a count is short, extraction missed a line."

    headers = ["Bank", "Lines extracted", "Tagged", "Still blank", "Flagged with a question"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    banks = list(SLICE.keys())
    tag_range = f"'{TAG_SHEET_NAME}'!$A$5:$A${last_data_row}"
    tag_g_range = f"'{TAG_SHEET_NAME}'!$G$5:$G${last_data_row}"
    tag_j_range = f"'{TAG_SHEET_NAME}'!$J$5:$J${last_data_row}"

    r = 5
    for bank in banks:
        ws.cell(row=r, column=1, value=bank).font = Font(name="Arial", bold=True)
        ws.cell(row=r, column=2, value=f"=COUNTIF({tag_range},A{r})")
        ws.cell(row=r, column=3, value=f'=COUNTIFS({tag_range},A{r},{tag_g_range},"<>")')
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=5, value=f'=COUNTIFS({tag_range},A{r},{tag_j_range},"<>")')
        r += 1

    total_row = r + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for col, letter in [(2, "B"), (3, "C"), (4, "D"), (5, "E")]:
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({letter}{r-len(banks)}:{letter}{r-1})")
        cell.font = Font(name="Arial", bold=True)
    return {}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate a concept-tagging Excel workbook from compiled_fs.db")
    parser.add_argument("--db", required=True, help="path to compiled_fs.db")
    parser.add_argument("--family", default=FAMILY, help=f"table family (only {FAMILY!r} supported this slice)")
    parser.add_argument("--out", default=None, help="output .xlsx path")
    args = parser.parse_args()

    out_path = Path(args.out) if args.out else Path("/tmp") / "tag_workbook_income_statement_fy25.xlsx"

    wb, summary = build_workbook(Path(args.db), family=args.family)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    print(f"Wrote {out_path}")
    for bank, counts in summary["per_bank"].items():
        print(f"  {bank}: {counts['rows']} rows, {counts['tagged']} tagged")
    if summary["warnings"]:
        print("Warnings:")
        for w in summary["warnings"]:
            print(f"  - {w}")
    else:
        print("Warnings: none")


if __name__ == "__main__":
    main()
