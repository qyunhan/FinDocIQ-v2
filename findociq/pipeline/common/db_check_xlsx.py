"""db_check_xlsx.py — STEP 4 verification view generated FROM the loaded
schema_v7 cells (NOT from the extraction JSON). Renders what the DB actually
holds so a human can eyeball it against the printed page: one sheet per
table_t row — rows in printed order, one column per col_period, hierarchy
shown by indentation, totals bold, verified sums_to annotated.

Run: python3 findociq/pipeline/common/db_check_xlsx.py --db findociq/db/compiled_fs.db
Out: findociq/outputs/checks/<dbname>.xlsx  (override with --out)
"""
import argparse
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parents[3]        # pipeline -> findociq -> repo


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db", default=str(REPO / "findociq" / "db" / "compiled_fs.db"))
    ap.add_argument("--out", default=None,
                    help="output xlsx; default findociq/outputs/checks/<dbname>.xlsx")
    args = ap.parse_args()
    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    tables = con.execute(
        "SELECT doc_id, table_id, table_title, section_id, page_range FROM table_t "
        "ORDER BY doc_id, table_id").fetchall()
    for t in tables:
        cols = con.execute(
            "SELECT col_id, col_period FROM col_dim "
            "WHERE doc_id=? AND table_id=? AND col_hierarchy=1 ORDER BY col_id",
            (t["doc_id"], t["table_id"])).fetchall()
        rows = con.execute(
            "SELECT row_id, row_hierarchy, row_parent, sums_to, row_leaf_label, line_no "
            "FROM row_dim WHERE doc_id=? AND table_id=? ORDER BY row_id",
            (t["doc_id"], t["table_id"])).fetchall()
        cells = {(c["row_id"], c["col_id"]): c for c in con.execute(
            "SELECT row_id, col_id, value_raw, cell_state FROM cell_fact "
            "WHERE doc_id=? AND table_id=?", (t["doc_id"], t["table_id"]))}

        ws = wb.create_sheet((t["table_id"][:28] or "t") + "…" if len(t["table_id"]) > 31
                             else t["table_id"])
        ws.append([f"{t['table_title']}  |  section {t['section_id']}  |  p{t['page_range']}"
                   f"  |  FROM DB {Path(args.db).name}"])
        ws["A1"].font = Font(bold=True)
        hdr = ["row", "lvl", "sums_to", "label"] + [c["col_period"] or f"col{c['col_id']}" for c in cols]
        ws.append(hdr)
        for cell in ws[2]:
            cell.font = Font(bold=True)
        for r in rows:
            is_total = r["sums_to"] is None and r["row_hierarchy"] == 0 and any(
                (r["row_id"], c["col_id"]) in cells for c in cols)
            printed_no = f"{r['line_no']}  " if r["line_no"] else ""
            label = ("    " * max(0, r["row_hierarchy"] - 1)) + printed_no + r["row_leaf_label"]
            line = [r["row_id"], r["row_hierarchy"],
                    r["sums_to"] if r["sums_to"] is not None else "",
                    label]
            for c in cols:
                cf = cells.get((r["row_id"], c["col_id"]))
                line.append(cf["value_raw"] if cf else "")
            ws.append(line)
            if is_total:
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
        ws.column_dimensions["D"].width = 46
    out = (Path(args.out).resolve() if args.out
           else REPO / "findociq" / "outputs" / "checks" / f"{Path(args.db).stem}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"{len(tables)} table sheet(s) -> {out}")


if __name__ == "__main__":
    main()
