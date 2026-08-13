import os
from pathlib import Path

import openpyxl
import pytest

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # pipeline/ on path
from common import tag_workbook as tw

DB = str(Path(__file__).resolve().parents[2] / "db" / "compiled_fs.db")


# ---------------------------------------------------------------------------
# derive_agg_and_groups
# ---------------------------------------------------------------------------

def test_derive_agg_and_groups_component_and_total():
    # rows 1, 2 sum into row 3; row 3 is itself not a component of anything.
    rows = [(1, 3), (2, 3), (3, None)]
    result = tw.derive_agg_and_groups(rows)
    assert result[1] == ("component", "G1")
    assert result[2] == ("component", "G1")
    assert result[3] == ("total", "G1")


def test_derive_agg_and_groups_atomic_standalone():
    rows = [(1, 3), (2, 3), (3, None), (4, None)]
    result = tw.derive_agg_and_groups(rows)
    assert result[4] == ("atomic", None)


def test_derive_agg_and_groups_multiple_groups_first_appearance_order():
    # row 1,2 -> target 5 (first group seen); row 3,4 -> target 6 (second group)
    rows = [(1, 5), (2, 5), (3, 6), (4, 6), (5, None), (6, None)]
    result = tw.derive_agg_and_groups(rows)
    assert result[1] == ("component", "G1")
    assert result[5] == ("total", "G1")
    assert result[3] == ("component", "G2")
    assert result[6] == ("total", "G2")


# ---------------------------------------------------------------------------
# is_noise_row
# ---------------------------------------------------------------------------

def test_is_noise_row_blank_and_footnote_and_numeric_only():
    assert tw.is_noise_row(None) is True
    assert tw.is_noise_row("") is True
    assert tw.is_noise_row("   ") is True
    assert tw.is_noise_row("Notes:") is True
    assert tw.is_noise_row("notes: see below") is True
    assert tw.is_noise_row("1,234 (56)") is True
    assert tw.is_noise_row("Net interest income") is False
    assert tw.is_noise_row("1 Unaudited.") is False  # has letters -> keep


# ---------------------------------------------------------------------------
# select_fy_column
# ---------------------------------------------------------------------------

def test_select_fy_column_picks_year_2025_over_distractors():
    cols = [
        {"col_id": 1, "col_leaf_label": "1st Half 2025", "col_period": "2025-06-30", "period_span": "1H", "unit": None},
        {"col_id": 2, "col_leaf_label": "% chg", "col_period": None, "period_span": None, "unit": "%"},
        {"col_id": 3, "col_leaf_label": "Year 2025", "col_period": None, "period_span": None, "unit": None},
        {"col_id": 4, "col_leaf_label": "Year 2024", "col_period": None, "period_span": None, "unit": None},
    ]
    chosen, warning = tw.select_fy_column(cols)
    assert chosen["col_id"] == 3
    assert warning is None


def test_select_fy_column_uses_period_span_fy_when_label_ambiguous():
    cols = [
        {"col_id": 1, "col_leaf_label": "2025", "col_period": "2025-12-31", "period_span": "FY", "unit": None},
        {"col_id": 2, "col_leaf_label": "2024", "col_period": "2024-12-31", "period_span": "FY", "unit": None},
        {"col_id": 100, "col_leaf_label": "GROUP", "col_period": None, "period_span": None, "unit": None},
    ]
    chosen, warning = tw.select_fy_column(cols)
    assert chosen["col_id"] == 1
    assert warning is None


def test_select_fy_column_none_found_returns_warning():
    cols = [
        {"col_id": 1, "col_leaf_label": "1st Half 2025", "col_period": None, "period_span": "1H", "unit": None},
    ]
    chosen, warning = tw.select_fy_column(cols)
    assert chosen is None
    assert warning is not None


# ---------------------------------------------------------------------------
# Light integration test against the real DB (guarded)
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not os.path.exists(DB), reason="compiled_fs.db not present")
def test_build_workbook_integration(tmp_path):
    wb, summary = tw.build_workbook(Path(DB))
    out = tmp_path / "wb.xlsx"
    wb.save(str(out))

    reopened = openpyxl.load_workbook(str(out))
    assert set(reopened.sheetnames) == {
        "START HERE", tw.TAG_SHEET_NAME, "Concept Dictionary", "Coverage check",
    }

    ws = reopened[tw.TAG_SHEET_NAME]
    header_vals = [c.value for c in ws[4]]
    assert header_vals == tw.TAG_HEADERS

    assert len(ws.data_validations.dataValidation) >= 1

    dbs_rows = [row for row in ws.iter_rows(min_row=5) if row[0].value == "DBS"]
    assert len(dbs_rows) > 0

    assert summary["per_bank"]["DBS"]["rows"] > 0
