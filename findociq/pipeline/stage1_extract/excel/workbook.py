"""stage1_extract.excel.workbook — openpyxl writing, cost sheets, API log, index/contents."""
from __future__ import annotations
import os, json, re, datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

import stage1_extract.chunk.schema as schema
from stage1_extract.chunk.schema import (
    GCell,
    INSTITUTION, DOC_TITLE, DOC_DATE, BRAND_COLOUR, HEADER_FILL,
    DARK_GREY, MID_GREY, WHITE, LIGHT_GREY, NUM_FMT,
    META_HEADERS, N_META,
    INPUT_PRICE_PER_M, OUTPUT_PRICE_PER_M, THINK_PRICE_PER_M,
    MODEL, INDEX_PATH, USAGE_LOG_PATH,
)


def coerce(v):
    """PDF text -> typed cell value."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—", "n.m.", "nm", "NA", "N/A"):
        return s
    if s.endswith("%"):
        return s
    t = s.replace(",", "")
    neg = t.startswith("(") and t.endswith(")")
    core = t[1:-1] if neg else t
    try:
        num = float(core)
    except ValueError:
        return s
    if neg:
        num = -num
    return int(num) if num == int(num) else num


def _hdr_style(cell, meta=False):
    cell.fill = PatternFill("solid", fgColor=schema.DARK_GREY if meta else schema.HEADER_FILL)
    cell.font = Font(bold=True, color=schema.WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table(ws, start_row: int, t) -> int:
    """Write one table (headers + data) starting at start_row. Returns the next free row."""
    nbase = N_META
    cols = t.columns
    ncols = nbase + len(cols)
    r = start_row

    base_headers = META_HEADERS[:-1] + [t.label_header or META_HEADERS[-1]]
    has_group = any(c.group for c in cols)

    if has_group:
        group_row, leaf_row = r, r + 1
        j, col = 0, nbase + 1
        while j < len(cols):
            g = cols[j].group or ""
            k = j
            while k + 1 < len(cols) and (cols[k + 1].group or "") == g:
                k += 1
            if g:
                gc = ws.cell(group_row, col, g)
                _hdr_style(gc)
                if k > j:
                    ws.merge_cells(start_row=group_row, start_column=col,
                                   end_row=group_row, end_column=col + (k - j))
            col += (k - j + 1)
            j = k + 1
        for ci, h in enumerate(base_headers):
            _hdr_style(ws.cell(leaf_row, 1 + ci, h), meta=(ci < 3))
        for ci, c2 in enumerate(cols):
            _hdr_style(ws.cell(leaf_row, nbase + 1 + ci, c2.leaf))
        r = leaf_row + 1
    else:
        for ci, h in enumerate(base_headers):
            _hdr_style(ws.cell(r, 1 + ci, h), meta=(ci < 3))
        for ci, c2 in enumerate(cols):
            _hdr_style(ws.cell(r, nbase + 1 + ci, c2.leaf))
        r += 1

    meta_font = Font(color=schema.MID_GREY, size=9)
    for row in t.rows:
        is_header = row.row_type == "section_header"
        is_total  = row.row_type == "total"
        is_note   = row.row_type == "note"

        ws.cell(r, 1, row.row_id)
        ws.cell(r, 2, row.level)
        ws.cell(r, 3, row.parent or "")
        indent = "    " * max(0, row.level - 1) if not is_header else ""
        ws.cell(r, 4, indent + row.label)
        col_cursor = nbase + 1
        for gcell in row.values:
            if isinstance(gcell, str):
                gcell = GCell.from_str(gcell)
            cell = ws.cell(r, col_cursor, coerce(gcell.value))
            state = gcell.cell_state
            if state == "grey":
                cell.fill = PatternFill("solid", fgColor=schema.LIGHT_GREY)
            elif state == "nil":
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif state == "zero":
                cell.value = 0
                cell.number_format = schema.NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            elif state == "reported" and isinstance(cell.value, (int, float)):
                cell.number_format = schema.NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            col_cursor += 1

        for ci in range(1, 4):
            ws.cell(r, ci).font = meta_font
            ws.cell(r, ci).alignment = Alignment(horizontal="center")
        if is_header:
            for ci in range(1, ncols + 1):
                cell = ws.cell(r, ci)
                cell.fill = PatternFill("solid", fgColor=schema.DARK_GREY)
                cell.font = Font(bold=True, color=schema.WHITE, size=10)
        elif is_total:
            for ci in range(1, ncols + 1):
                cur = ws.cell(r, ci).font
                ws.cell(r, ci).font = Font(bold=True, color=cur.color, size=cur.size or 10)
        elif is_note:
            for ci in range(1, ncols + 1):
                ws.cell(r, ci).font = Font(italic=True, color=schema.MID_GREY, size=8)
        r += 1
    return r + 1


def write_cost_sheet(wb, call_log: list[dict], run_usage: dict, out_path: str):
    """Write (or replace) a 'Cost' tab summarising every API call made this run."""
    if "Cost" in wb.sheetnames:
        wb.remove(wb["Cost"])
    ws = wb.create_sheet("Cost")

    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, f"{schema.INSTITUTION}  |  API Cost Log  |  {schema.DOC_DATE}")
    c.fill = PatternFill("solid", fgColor=schema.BRAND_COLOUR)
    c.font = Font(bold=True, color=schema.WHITE, size=12)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["#", "Timestamp", "Label", "Model", "Image?",
               "Input tok", "Output tok", "Think tok", "Est. cost (USD)"]
    for ci, h in enumerate(headers, 1):
        _hdr_style(ws.cell(2, ci, h))

    total_cost = 0.0
    for ri, rec in enumerate(call_log, start=1):
        cost = rec.get("est_cost_usd", 0) or 0
        total_cost += cost
        ws.cell(ri + 2, 1, ri).alignment = Alignment(horizontal="center")
        ws.cell(ri + 2, 2, rec.get("ts", ""))
        ws.cell(ri + 2, 3, rec.get("label", ""))
        ws.cell(ri + 2, 4, rec.get("model", schema.MODEL))
        ws.cell(ri + 2, 5, "yes" if rec.get("image_used") else "no").alignment = Alignment(horizontal="center")
        for ci, key in enumerate(["prompt_tokens", "output_tokens", "thinking_tokens"], start=6):
            cell = ws.cell(ri + 2, ci, rec.get(key, 0))
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        cell = ws.cell(ri + 2, 9, round(cost, 5))
        cell.number_format = "$#,##0.00000"
        cell.alignment = Alignment(horizontal="right")

    sr = len(call_log) + 4
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
    c = ws.cell(sr, 1, f"TOTAL  ({run_usage['calls']} calls)")
    c.font = Font(bold=True, size=10)
    for ci, key in enumerate(["prompt", "output", "thinking"], start=6):
        cell = ws.cell(sr, ci, run_usage.get(key, 0))
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
    cell = ws.cell(sr, 9, round(run_usage.get("cost", 0), 5))
    cell.number_format = "$#,##0.00000"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right")

    ws.merge_cells(start_row=sr + 1, start_column=1, end_row=sr + 1, end_column=9)
    note = ws.cell(sr + 1, 1,
        f"Pricing: ${schema.INPUT_PRICE_PER_M}/M input, ${schema.OUTPUT_PRICE_PER_M}/M output, "
        f"${schema.THINK_PRICE_PER_M}/M thinking  "
        f"|  Model: {schema.MODEL}  |  Log: {os.path.basename(schema.USAGE_LOG_PATH)}")
    note.font = Font(italic=True, color=schema.MID_GREY, size=8)

    for col, w in {1:4, 2:20, 3:28, 4:18, 5:8, 6:12, 7:12, 8:12, 9:16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def save_cost_summary(call_log: list[dict], run_usage: dict, out_path: str,
                      summary_path: str = "", document: str = ""):
    """Write/merge a JSON cost summary to logs/cost_summary.json.

    Two doc_ids sharing a bank_period run directory (e.g. DBS_1Q22_pillar3 and
    DBS_1Q22_trading_update both write to outputs/pillar3/dbs_1Q22/) used to
    silently overwrite each other's cost record here, since this wrote a
    single flat "calls" list for whichever doc ran last. Fixed by (1) stamping
    `document` on every call (mirroring append_to_api_log, which already did
    this), and (2) MERGING into any existing summary_path instead of
    overwriting, de-duplicated on (document, ts, label) so re-running one doc
    doesn't double-count. `totals` stays the combined figure across every call
    ever logged to this file; `by_document` breaks the same calls out per doc
    so a per-doc cost/token figure survives sibling docs sharing the dir.
    """
    summary_path = summary_path or (os.path.splitext(out_path)[0] + "_cost_summary.json")
    for rec in call_log:
        rec.setdefault("document", document)

    existing_calls: list[dict] = []
    if os.path.exists(summary_path):
        try:
            existing_calls = json.load(open(summary_path)).get("calls", [])
        except Exception:
            existing_calls = []
    seen = {(c.get("document"), c.get("ts"), c.get("label")) for c in existing_calls}
    merged = existing_calls + [c for c in call_log
                               if (c.get("document"), c.get("ts"), c.get("label")) not in seen]

    def _totals(calls: list[dict]) -> dict:
        t = {"calls": len(calls), "input_tokens": 0, "output_tokens": 0,
             "thinking_tokens": 0, "est_cost_usd": 0.0}
        for c in calls:
            t["input_tokens"] += c.get("prompt_tokens", 0) or 0
            t["output_tokens"] += c.get("output_tokens", 0) or 0
            t["thinking_tokens"] += c.get("thinking_tokens", 0) or 0
            t["est_cost_usd"] += c.get("est_cost_usd", 0) or 0
        t["total_tokens"] = t["input_tokens"] + t["output_tokens"] + t["thinking_tokens"]
        t["est_cost_usd"] = round(t["est_cost_usd"], 5)
        return t

    by_document: dict[str, dict] = {}
    for c in merged:
        by_document.setdefault(c.get("document") or "(unknown)", []).append(c)
    by_document = {doc: _totals(calls) for doc, calls in by_document.items()}

    summary = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "output_file": os.path.basename(out_path),
        "model": schema.MODEL,
        "pricing": {"input_per_million": schema.INPUT_PRICE_PER_M,
                    "output_per_million": schema.OUTPUT_PRICE_PER_M},
        "totals": _totals(merged),
        "by_document": by_document,
        "calls": merged,
    }
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"   💰 cost summary → {summary_path}")


def _describe_call(label: str, image_used: bool, bank: str) -> str:
    """One-line human description of what a call was doing."""
    parts = label.split("_")
    is_chunk = label.endswith(("_c1", "_c2", "_c3", "_c4", "_c5"))
    is_multi = "multi" in label
    page_tok = next((p for p in parts if p.startswith("p") and any(c.isdigit() for c in p)), "")
    section_parts = [p for p in parts if p not in (page_tok.lstrip("p"),) and not p.startswith("p")]
    section_id = ".".join(section_parts[:-1]) if is_chunk else ".".join(section_parts)
    section_id = section_id.strip("._")
    desc = f"{bank} §{section_id} {page_tok}"
    if is_multi:
        desc += " (shared page — multiple sections)"
    if is_chunk:
        chunk_n = label.rsplit("_c", 1)[-1]
        desc += f" chunk {chunk_n}"
    if image_used:
        desc += " +image"
    return desc.strip()


_API_LOG_HEADERS = [
    "#", "Run date", "Bank", "Section label", "Description",
    "Model", "Pages", "Image?",
    "Input tok", "Output tok", "Think tok", "Total tok",
    "Est. cost (USD)", "Cumulative cost (USD)",
]


def append_to_api_log(call_log: list[dict], bank: str, log_path: str,
                      period: str = "", document: str = ""):
    """Append this run's calls to the api_log.xlsx for this run."""
    for rec in call_log:
        if "document" not in rec and document:
            rec["document"] = document
    if os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
        last_row = ws.max_row
        cum_cost = 0.0
        for r in range(2, last_row + 1):
            v = ws.cell(r, 14).value
            if isinstance(v, (int, float)):
                cum_cost = v
        next_row = last_row + 1
        next_num = last_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "API Calls"
        for ci, h in enumerate(_API_LOG_HEADERS, 1):
            c = ws.cell(1, ci, h)
            c.fill = PatternFill("solid", fgColor=schema.HEADER_FILL)
            c.font = Font(bold=True, color=schema.WHITE, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        next_row, next_num, cum_cost = 2, 1, 0.0

    run_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for rec in call_log:
        label      = rec.get("label", "")
        image_used = rec.get("image_used", False)
        prompt_t   = rec.get("prompt_tokens", 0) or 0
        output_t   = rec.get("output_tokens", 0) or 0
        think_t    = rec.get("thinking_tokens", 0) or 0
        cost       = rec.get("est_cost_usd", 0) or 0
        cum_cost  += cost
        desc = _describe_call(label, image_used, bank)
        page_hint = next((p for p in label.split("_") if p.startswith("p") and any(c.isdigit() for c in p)), "")

        row_vals = [
            next_num, run_date, bank, label, desc,
            rec.get("model", schema.MODEL), page_hint,
            "yes" if image_used else "no",
            prompt_t, output_t, think_t, prompt_t + output_t + think_t,
            round(cost, 5), round(cum_cost, 5),
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(next_row, ci, v)
            if ci in (9, 10, 11, 12):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif ci in (13, 14):
                cell.number_format = "$#,##0.00000"
                cell.alignment = Alignment(horizontal="right")
            elif ci == 8:
                cell.alignment = Alignment(horizontal="center")
        next_row += 1
        next_num += 1

    for col, w in {1:5, 2:18, 3:8, 4:26, 5:48, 6:20, 7:12, 8:8,
                   9:13, 10:13, 11:13, 12:13, 13:16, 14:18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    wb.save(log_path)
    print(f"   📋 API log updated → {log_path}  (total log rows: {next_row - 2})")


def write_section_header(ws, section_id: str, title: str, last_col: int,
                         table_label: str = "", table_n: int = 0, total_tables: int = 0):
    """Row 1 = brand banner with section title; row 2 = table identifier; row 3 = source line."""
    last_col = max(last_col, N_META)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, f"{schema.INSTITUTION}  |  Section {section_id}: {title}")
    c.fill = PatternFill("solid", fgColor=schema.BRAND_COLOUR)
    c.font = Font(bold=True, color=schema.WHITE, size=12)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ref = f"[{section_id} Table {table_n}]" if table_n else f"[{section_id}]"
    count_str = f"  ({table_n} of {total_tables})" if total_tables > 1 else ""
    label_str = f"  —  {table_label}" if table_label else ""
    c = ws.cell(2, 1, f"{ref}{count_str}{label_str}")
    c.font = Font(bold=True, color=schema.BRAND_COLOUR, size=10)
    ws.row_dimensions[2].height = 16

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(3, 1, f"Source: {schema.DOC_TITLE}, {schema.DOC_DATE}  |  Units: S$ millions unless noted")
    c.font = Font(italic=True, color=schema.MID_GREY, size=9)
    ws.row_dimensions[3].height = 14


def style_sheet_columns(ws):
    widths = {1: 13, 2: 14, 3: 14, 4: 58}
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 15)


def sheet_name(used: set, section_id: str, title: str) -> str:
    title = re.sub(r'[\x00-\x1f\x7f]', ' ', title)
    base = f"{section_id} - {title}".strip(" -")
    for ch in '[]:*?/\\':
        base = base.replace(ch, " ")
    base = " ".join(base.split())[:31].rstrip() or section_id[:31].rstrip() or "Sheet"
    name, i = base, 2
    while name in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)].rstrip() + suffix
        i += 1
    used.add(name)
    return name


def table_sheet_name(used: set, section_id: str, table_n: int, total_tables: int = 1) -> str:
    """One tab per table. A section with MULTIPLE tables gets a ' Table N'
    suffix to disambiguate ('18.4 Table 1', '18.4 Table 2'); a section with a
    single table keeps just its (truncated) section id — no gratuitous
    ' Table 1' eating into the visible section name. This mirrors
    write_section_header, which only prints '(N of M)' when total_tables > 1.

    Truncate the section_id BEFORE appending any suffix (not the concatenated
    string afterwards) so a long section_id never amputates the suffix — a bare
    `[:31]` on `f"{section_id} Table {table_n}"` was cutting into 'Table N'
    (e.g. 'key_financial_ratios_2_3 Table ' lost the trailing '1'). rstrip()
    after every truncation so no name ends in a dangling space.
    """
    sid = section_id
    for ch in '[]:*?/\\':
        sid = sid.replace(ch, " ")
    sid = " ".join(sid.split())
    suffix = f" Table {table_n}" if total_tables > 1 else ""
    sid = sid[:max(31 - len(suffix), 0)].rstrip()
    base = f"{sid}{suffix}".strip() or f"t{table_n}"
    base = base[:31].rstrip()
    name, i = base, 2
    while name in used:
        coll_suffix = f" ({i})"
        name = base[:31 - len(coll_suffix)].rstrip() + coll_suffix
        i += 1
    used.add(name)
    return name


def load_index() -> list[dict]:
    if os.path.exists(schema.INDEX_PATH):
        try:
            idx = json.load(open(schema.INDEX_PATH))
            for e in idx:
                if "sheet" not in e and "first_tab" in e:
                    e["sheet"] = e["first_tab"]
            return idx
        except Exception:
            return []
    return []


def save_index(idx: list[dict]):
    os.makedirs(os.path.dirname(schema.INDEX_PATH) or ".", exist_ok=True)
    json.dump(idx, open(schema.INDEX_PATH, "w"), indent=2)


def update_index(idx: list[dict], entry: dict) -> list[dict]:
    idx = [e for e in idx if e["section_id"] != entry["section_id"]]
    idx.append(entry)
    return idx


def rebuild_contents(wb, idx: list[dict]):
    """(Re)build the Contents sheet as the first tab, hyperlinked to each section."""
    if "Contents" in wb.sheetnames:
        wb.remove(wb["Contents"])
    ws = wb.create_sheet("Contents", 0)

    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, schema.INSTITUTION)
    c.fill = PatternFill("solid", fgColor=schema.BRAND_COLOUR)
    c.font = Font(bold=True, color=schema.WHITE, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1, f"{schema.DOC_TITLE} — {schema.DOC_DATE}")
    c.font = Font(italic=True, size=11)
    c.alignment = Alignment(horizontal="center")

    headers = ["Section", "Title", "Pages", "Tables", "Sheet"]
    for ci, h in enumerate(headers, start=1):
        _hdr_style(ws.cell(4, ci, h))

    r = 5
    for e in sorted(idx, key=lambda x: x.get("first_page", 0)):
        ws.cell(r, 1, e["section_id"]).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, e["title"])
        ws.cell(r, 3, e.get("pages", "")).alignment = Alignment(horizontal="center")
        ws.cell(r, 4, e.get("n_tables", "")).alignment = Alignment(horizontal="center")
        sname = e.get("sheet") or e.get("first_tab", "")
        link = ws.cell(r, 5, sname)
        if sname and sname in wb.sheetnames:
            link.hyperlink = f"#'{sname}'!A1"
        link.font = Font(color="0000CC", underline="single")
        r += 1

    for col, w in {1: 12, 2: 60, 3: 16, 4: 10, 5: 32}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return ws
