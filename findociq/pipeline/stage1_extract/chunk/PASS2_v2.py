"""PASS2_v2.py — thin entrypoint using the pass2/ package.

PASS2_Extract_to_Excel.py is the frozen monolith fallback.
This file contains ONLY argparse + main(); all logic lives in pass2/.
"""
from __future__ import annotations
import os, sys, re, argparse, threading
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

# run_doc launches this file as a SCRIPT with cwd=stage1_extract/chunk, so sys.path[0]
# is chunk/ — not pipeline/ — and every `stage1_extract.*` / `stage2_load.*` import
# below would fail. Put pipeline/ on the path first (same idiom as toc/toc_stage.py:44).
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))   # chunk -> stage1_extract -> pipeline

from stage1_extract.gemini.gemini_client import build_client  # noqa: E402  (pipeline/gemini_client.py)

# All logic from the pass2 package — nothing redefined here.
import stage1_extract.chunk.schema as schema
from stage1_extract.chunk.schema import (
    BANKS, GTable,
    _run_usage,          # read-only references to the single package instances
    _call_log,           # these are never re-assigned in this file
    _call_log_lock,
    _run_usage_lock,
    _pdfium_lock,
)
from stage1_extract.chunk.render import detect_bank, derive_period, page_is_narrative, page_has_table_structure
from stage1_extract.chunk.transforms import (
    validate_numbers, validate_exactly_once, _apply_transforms,
    drop_misowned_tables, flag_duplicate_tables, is_true_continuation,
)
from stage1_extract.chunk.extract import (
    build_units, load_sections, group_key,
    load_cached_unit, extract_unit_chunked,
    route_tables,
    _is_transport_error, _PROMPT_HASH,
    _migrate_legacy_chunk_dirs,
)
from stage1_extract.chunk.batch import run_batch
from stage1_extract.excel.workbook import (
    load_index, save_index, update_index, rebuild_contents,
    write_table, write_cost_sheet, save_cost_summary, append_to_api_log,
    write_section_header, style_sheet_columns,
    table_sheet_name,
)


def bucket_tables(bucket, tables) -> None:
    """Append `tables` to `bucket`, merging genuine page-break continuations.

    `continued_from_previous` is the MODEL's claim. schema.GTable defines it as
    "rows continue under the same columns, header NOT repeated" — so a table
    that carries its own title, or whose first substantive row is a section /
    sub header, contradicts the claim and is a NEW TABLE that merely prints
    '(cont'd)' in its caption. UOB 2Q26 page 6 is exactly that: title
    "Financial Highlights (cont'd)", first row the section header "Key
    financial ratios (%) (cont'd)", and the TOC had already split pages 5 and 6
    into separate units so there was nothing in the bucket to merge into anyway.

    THE FLAG IS CLEARED WHEN THE MERGE DOES NOT FIRE. Both branches already
    recorded such a table correctly — `bucket.append(t)` — but left the boolean
    set, and `load_v7._load_table` refuses any table still carrying it (its
    contract is that continuations are merged before loading). The loader was
    being handed a contradiction: 'I am a fragment' on a self-describing table.
    One stale boolean aborted the ENTIRE UOB 2Q26 load — 47 units, 832 verified
    rows — so the flag must never outlive the decision it feeds.

    One implementation, called from every bucketing site. There were two
    divergent copies: the cached-unit path tested 3 conditions and the live path
    5, so an identical table merged or did not depending purely on whether its
    unit happened to be cached.
    """
    for t in tables:
        if is_true_continuation(bucket[-1] if bucket else None, t):
            bucket[-1].rows.extend(t.rows)
        else:
            t.continued_from_previous = False
            bucket.append(t)


def main():
    ap = argparse.ArgumentParser(description="Simplified PDF -> Excel extraction (one tab per section)")
    ap.add_argument("pdf")
    ap.add_argument("--toc", default=None, help="path to TOC JSON (default: out/<bank>_toc.json or out/step1_toc.json)")
    ap.add_argument("--section", help="only this section_id")
    ap.add_argument("--start-section", help="begin at this section_id (document order)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--no-pause", action="store_true", help="do not pause after each section")
    ap.add_argument("--image", action="store_true", help="always attach a rendered image alongside the PDF")
    ap.add_argument("--thinking", action="store_true", help="enable model thinking (higher cost)")
    ap.add_argument("--force", action="store_true", help="re-extract sections whose tab already exists")
    ap.add_argument("--list", action="store_true", help="list sections and exit")
    ap.add_argument("--no-audit", action="store_true",
                    help="skip writing audit files (prompt.txt, pages.pdf, response.txt, parsed.json, meta.json)")
    ap.add_argument("--dry-run", action="store_true",
                    help="print the per-section call plan (units/prompts/tabs) and exit — no API, no key")
    ap.add_argument("--bank", choices=list(BANKS), help="force the institution/brand (else auto-detected)")
    ap.add_argument("--institution", help="override the banner institution name")
    ap.add_argument("--brand", help="override the brand colour hex (e.g. 1B6EC2)")
    ap.add_argument("--doc-date", help="override the source-line date (e.g. '31 December 2025')")
    ap.add_argument("--chunk-pages", type=int, default=2, metavar="N",
                    help="max pages per Gemini call for spanning sections (default 2; use 0 to disable chunking)")
    ap.add_argument("--workers", type=int, default=5, metavar="N",
                    help="max concurrent section groups (default 5; use 1 for sequential)")
    ap.add_argument("--batch", action="store_true",
                    help="run extraction via the Gemini BATCH API (async, 50%% cost). "
                         "Pure billing/execution mode — prompts/config/contract byte-identical "
                         "to the sync path. Submits work in dependency rounds instead of the thread pool.")
    ap.add_argument("--batch-timeout", type=int, default=1800, metavar="SEC",
                    help="hard cap for polling a batch job to terminal state (default 1800s / 30min)")
    ap.add_argument("--out-root", default=None, metavar="DIR",
                    help="override the outputs/pillar3 root for ALL run paths "
                         "(audit, ledger, cost, xlsx) — used to redirect smoke runs to a scratch dir")
    ap.add_argument("--family", default=None,
                    help="doc family from the router (classify/family.py), e.g. pillar3|fs — "
                         "decides the output root/filenames and DOC_TITLE. Absent -> PASS2 "
                         "self-classifies the PDF; never silently assumed to be pillar3.")
    args = ap.parse_args()

    if not os.path.exists(args.pdf):
        sys.exit(f"PDF not found: {args.pdf}")

    detected, det_date = detect_bank(args.pdf)
    bank = args.bank or detected

    bank_slug = (bank or "dbs").lower()
    doc_stem  = Path(args.pdf).stem

    _early_date = args.doc_date or det_date or schema.DOC_DATE
    period = derive_period(_early_date, doc_stem)

    # --family: the router (run_doc.py) passes the family it already classified,
    # keeping it authoritative and visible on the command line. Standalone runs
    # self-classify via classify/family.py rather than assuming pillar3 (see
    # docs/specs/2026-07-29-family-aware-output-paths.md).
    if args.family:
        _family_raw = args.family
    else:
        try:
            from stage1_extract.route.family import classify as _classify_doc
            _family_raw = _classify_doc(args.pdf).get("family")
        except Exception as e:                                   # noqa: BLE001
            print(f"   ⚠️  self-classification failed ({type(e).__name__}: {e})")
            _family_raw = None
    family, schema.DOC_TITLE, _family_known = schema.resolve_family(_family_raw)

    if args.out_root:
        # --out-root redirects EVERY family, not just pillar3. pillar3 resolves
        # through _P3_ROOT (byte-identical legacy behaviour); non-pillar3
        # families resolve through _OUTPUTS_ROOT/<family> — so both must point
        # at the override, else fs/... silently writes into the repo working
        # tree (findociq/outputs/fs/...) instead of the requested dir.
        schema._P3_ROOT = Path(args.out_root)
        schema._OUTPUTS_ROOT = Path(args.out_root)
        print(f"   ↪ out-root override: all run paths under {Path(args.out_root)}")

    rp = schema.RunPaths(bank_slug, period, doc_stem, family=family)
    rp.makedirs()

    # Resolve TOC path — write to schema so load_sections() picks it up
    if args.toc:
        schema.TOC_PATH = args.toc
    elif rp.toc.exists():
        schema.TOC_PATH = str(rp.toc)
    else:
        _flat_keyed  = schema._P3_ROOT / f"{bank_slug}_{period}_toc.json"
        _flat_stem   = schema._P3_ROOT / f"{doc_stem}_toc.json"
        _legacy      = schema._P3_ROOT / "step1_toc.json"
        if _flat_keyed.exists():
            schema.TOC_PATH = str(_flat_keyed)
        elif _flat_stem.exists():
            schema.TOC_PATH = str(_flat_stem)
        else:
            if _legacy.exists():
                print(f"   ℹ️  legacy step1_toc.json found but ignored — it may be for a different document")
            sys.exit(
                f"No TOC found for '{args.pdf}'.\n"
                f"Expected: {rp.toc}\n"
                f"Run: python3 DELIVERABLE/pillar3/PASS1_TOC.py \"{args.pdf}\""
            )

    out_path = args.out or str(rp.xlsx)

    schema.INDEX_PATH     = str(rp.index)
    schema.USAGE_LOG_PATH = str(rp.ledger)
    schema.COST_LOG_PATH  = str(rp.cost)
    schema.AUDIT_DIR      = str(rp.audit_dir)

    _migrate_legacy_chunk_dirs(schema.AUDIT_DIR)

    if bank:
        schema.INSTITUTION  = BANKS[bank]["institution"]
        schema.BRAND_COLOUR = BANKS[bank]["brand"]
    if det_date:
        schema.DOC_DATE = det_date
    if args.institution:
        schema.INSTITUTION = args.institution
    if args.brand:
        schema.BRAND_COLOUR = args.brand
    if args.doc_date:
        schema.DOC_DATE = args.doc_date
    print(f"🏦 Institution: {schema.INSTITUTION}  |  brand #{schema.BRAND_COLOUR}  |  date: {schema.DOC_DATE}"
          f"  ({'detected ' + bank if bank and not args.bank else ('--bank ' + bank if args.bank else 'default')})")
    print(f"📁 Family: {family}  |  title: {schema.DOC_TITLE}"
          f"  ({'--family' if args.family else 'self-classified'})")

    document, sections = load_sections()
    if args.list:
        print(f"{document.get('title','')[:60]} — {len(sections)} sections (document order):")
        for s in sections:
            print(f"  {s['section_id']:<8} p{s['start_page']}-{s['end_page']:<3} "
                  f"[{s.get('page_label') or '—'}]  {s['title'][:50]}")
        return

    current_doc = os.path.basename(args.pdf)
    if os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
        stored_doc = wb.properties.subject or ""
        if stored_doc and stored_doc != current_doc:
            sys.exit(
                f"\n❌ PROVENANCE MISMATCH — refusing to write.\n"
                f"   Workbook '{os.path.basename(out_path)}' was created from: {stored_doc}\n"
                f"   Current PDF is:                                           {current_doc}\n"
                f"   Use a different output path or --out to avoid cross-document contamination."
            )
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.properties.subject = current_doc
    used_names = set(wb.sheetnames)
    idx = load_index()

    units = build_units(sections)
    groups: dict[str, list[dict]] = {}
    for u in units:
        groups.setdefault(u["group"], []).append(u)

    if args.dry_run:
        sel = [(g, us) for g, us in groups.items()
               if not args.section or args.section == g
               or any(args.section == lf["section_id"] for u in us for lf in u["leaves"])]
        total = sum(len(us) for _, us in sel)
        print(f"DRY RUN — {len(sel)} section group(s), {total} Gemini call(s) planned (no API used):")
        for gnum, gunits in sel:
            print(f"\n##### Section {gnum}")
            for u in gunits:
                pr = "+".join(map(str, u["pages"]))
                tabs = ", ".join(lf["section_id"] for lf in u["leaves"])
                print(f"   [{u['type']:8}] pages {pr:<8} -> PROMPT_{u['type'].upper():8} -> tab(s): {tabs}")
        return

    if args.section:
        matched = any(
            args.section == g or
            any(args.section == lf["section_id"] for u in us for lf in u["leaves"])
            for g, us in groups.items()
        )
        if not matched:
            available = sorted({s["section_id"] for s in sections})
            sys.exit(f"No section matched '--section {args.section}'. "
                     f"Available: {', '.join(available)}")

    _client = None
    _client_lock = threading.Lock()

    def _get_client():
        nonlocal _client
        with _client_lock:
            if _client is None:
                _client = build_client(http_options={"timeout": 120_000})
        return _client

    def _tab_exists(sid: str) -> bool:
        return any(nm == sid or nm.startswith(f"{sid} -") or nm.startswith(f"{sid} Table ") for nm in wb.sheetnames)

    # In --batch mode this is populated up-front by run_batch: {unit_id: (ext, meta)}.
    # _extract_group reads it instead of making a sync API call. Units missing from
    # it (batch error / not selected) fall through to the normal sync path.
    precomputed: dict = {}

    def _extract_group(gnum: str, gunits: list, group_leaves: list,
                       leaf_target: str | None) -> dict[str, list]:
        grp_tables: dict[str, list] = defaultdict(list)
        for u in gunits:
            if leaf_target and not any(lf["section_id"] == leaf_target for lf in u["leaves"]):
                continue
            pr = "+".join(map(str, u["pages"]))
            leaf_ids = ", ".join(lf["section_id"] for lf in u["leaves"])
            if u["type"] in ("single", "multiple") and (
                    page_is_narrative(args.pdf, u["pages"][0]) or
                    not page_has_table_structure(args.pdf, u["pages"][0])):
                print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — narrative, skipped (no call)")
                continue
            audit_exists = os.path.exists(os.path.join(schema.AUDIT_DIR, u["unit_id"], "parsed.json"))
            if not args.force and not args.no_audit and audit_exists:
                ext = load_cached_unit(u, args)
                if ext is not None:
                    sids = tuple(lf["section_id"] for lf in u["leaves"])
                    num_issues = validate_numbers(ext, args.pdf, u["pages"], section_ids=sids, unit=u)
                    if num_issues:
                        print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — resumed from audit "
                              f"(⚠ {len(num_issues)} number issues in cache)")
                    else:
                        print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — resumed from audit (no call)")
                    tables, drop_records = drop_misowned_tables(_apply_transforms(ext.tables), u, args.pdf, u["pages"])
                    if drop_records and not args.no_audit:
                        import json as _json
                        mj = os.path.join(schema.AUDIT_DIR, u["unit_id"], "meta.json")
                        if os.path.exists(mj):
                            _m = _json.load(open(mj))
                            _m["table_drops"] = drop_records
                            _m["all_dropped"] = all(r["dropped"] for r in drop_records)
                            _json.dump(_m, open(mj, "w"), indent=2)
                    with _all_tables_lock:
                        _all_tables_by_unit[u["unit_id"]] = tables
                    if u["type"] == "multiple":
                        for t, lf, method, sc, flagged in route_tables(tables, u["leaves"]):
                            bucket_tables(grp_tables[lf["section_id"]], [t])
                    else:
                        bucket_tables(grp_tables[u["leaves"][0]["section_id"]], tables)
                    continue
            if args.batch and u["unit_id"] in precomputed:
                ext, meta = precomputed[u["unit_id"]]
                print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — from batch (no sync call)")
            else:
                print(f"   • [{u['type']:8}] p{pr:<7} prompt=PROMPT_{u['type'].upper():8} -> tab(s): {leaf_ids}")
                try:
                    chunk_size = args.chunk_pages if args.chunk_pages > 0 else 9999
                    ext, meta = extract_unit_chunked(_get_client(), args.pdf, u,
                                                      force_image=args.image,
                                                      with_thinking=args.thinking,
                                                      save_audit=not args.no_audit,
                                                      chunk_size=chunk_size)
                except Exception as e:
                    err_tag = "503/timeout" if _is_transport_error(e) else e.__class__.__name__
                    print(f"     ❌ FAILED ({err_tag}): {e}")
                    for lf in u["leaves"]:
                        grp_tables[lf["section_id"]].append("__FAILED__")
                    continue
            tables, drop_records = drop_misowned_tables(_apply_transforms(ext.tables), u, args.pdf, u["pages"])
            if drop_records and not args.no_audit:
                import json as _json
                mj = os.path.join(schema.AUDIT_DIR, u["unit_id"], "meta.json")
                if os.path.exists(mj):
                    _m = _json.load(open(mj))
                    _m["table_drops"] = drop_records
                    _m["all_dropped"] = all(r["dropped"] for r in drop_records)
                    _json.dump(_m, open(mj, "w"), indent=2)
            with _all_tables_lock:
                _all_tables_by_unit[u["unit_id"]] = tables
            if u["type"] == "multiple":
                for t, lf, method, sc, flagged in route_tables(tables, u["leaves"]):
                    bucket_tables(grp_tables[lf["section_id"]], [t])
                    mark = "⚠ " if flagged else "  "
                    print(f"        {mark}→ [{method:5} {sc:.2f}] '{t.title[:30]}'  →  tab {lf['section_id']}")
                if len(tables) != len({lf["section_id"] for lf in u["leaves"]}):
                    print(f"        ⚠ {len(tables)} table(s) vs {len(u['leaves'])} subsection(s) on this page")
            else:
                bucket_tables(grp_tables[u["leaves"][0]["section_id"]], tables)
                if u["type"] == "spanning":
                    print(f"        ({len(tables)} table(s) kept across pages {'+'.join(map(str,u['pages']))})")
            ut, tag = meta["usage"], (" +img" if meta["image_used"] else "")
            print(f"     ✓ {len(tables)} table(s){tag}  "
                  f"[{ut.get('prompt_tokens','?')}in/{ut.get('output_tokens','?')}out/"
                  f"{ut.get('thinking_tokens','?')}think tok]")
        return dict(grp_tables)

    _all_tables_by_unit: dict[str, list] = {}
    _all_tables_lock = threading.Lock()
    section_tables: dict[str, list[GTable]] = defaultdict(list)
    started = args.start_section is None

    ordered_groups: list[tuple] = []
    for gnum, gunits in groups.items():
        group_leaves = [s for s in sections if group_key(s) == gnum]
        unit_leaf_ids = {lf["section_id"] for u in gunits for lf in u["leaves"]}
        extra_leaves = [s for s in sections if s["section_id"] in unit_leaf_ids
                        and s not in group_leaves]
        group_leaves = group_leaves + extra_leaves
        leaf_ids_in_group = {s["section_id"] for s in group_leaves}

        if args.section and not (args.section == gnum or args.section in leaf_ids_in_group):
            continue
        if not started:
            if gnum == args.start_section:
                started = True
            else:
                continue
        if not args.force and not args.section and group_leaves and all(_tab_exists(s["section_id"]) for s in group_leaves):
            print(f"⏭️  group {gnum} already present — skip (use --force to redo)")
            continue

        leaf_target = args.section if (args.section and args.section in leaf_ids_in_group) else None
        ordered_groups.append((gnum, gunits, group_leaves, leaf_target))

    if args.force:
        if args.section:
            sids_to_clear = {args.section}
        else:
            sids_to_clear = {
                lf["section_id"]
                for _, _, group_leaves, _ in ordered_groups
                for lf in group_leaves
            }
        for sid in sids_to_clear:
            sid_slug = sid.replace(".", "_")
            audit_bank_dir = Path(schema.AUDIT_DIR)
            if audit_bank_dir.exists():
                for unit_dir in list(audit_bank_dir.iterdir()):
                    if unit_dir.is_dir() and unit_dir.name.startswith(sid_slug):
                        import shutil
                        shutil.rmtree(unit_dir)
                        print(f"   🗑  --force deleted audit '{unit_dir.name}'")
        for sname in list(wb.sheetnames):
            if sname in ("Contents", "Cost"):
                continue
            for sid in sids_to_clear:
                if (sname == sid or
                        sname.startswith(f"{sid} Table ") or
                        sname.startswith(f"{sid} -")):
                    wb.remove(wb[sname])
                    used_names.discard(sname)
                    print(f"   🗑  --force cleared tab '{sname}'")
                    break

    # ── BATCH MODE ──────────────────────────────────────────────────────────
    # Up-front, run every unit that WOULD make a sync API call through the Batch
    # API (dependency rounds). Selection mirrors _extract_group's gates so the
    # set is identical to what the sync path would call. Results feed precomputed;
    # the worker pool below then just reads them (no per-unit API calls).
    if args.batch:
        to_batch = []
        for gnum, gunits, group_leaves, leaf_target in ordered_groups:
            for u in gunits:
                if leaf_target and not any(lf["section_id"] == leaf_target for lf in u["leaves"]):
                    continue
                if u["type"] in ("single", "multiple") and (
                        page_is_narrative(args.pdf, u["pages"][0]) or
                        not page_has_table_structure(args.pdf, u["pages"][0])):
                    continue
                audit_exists = os.path.exists(
                    os.path.join(schema.AUDIT_DIR, u["unit_id"], "parsed.json"))
                if not args.force and not args.no_audit and audit_exists:
                    if load_cached_unit(u, args) is not None:
                        continue
                to_batch.append(u)
        if to_batch:
            chunk_size = args.chunk_pages if args.chunk_pages > 0 else 9999
            print(f"\n🧺 BATCH MODE — {len(to_batch)} unit(s) via Gemini Batch API "
                  f"(async, 50% cost; prompts/config identical to sync)")
            precomputed.update(
                run_batch(_get_client(), args.pdf, to_batch,
                          force_image=args.image, with_thinking=args.thinking,
                          save_audit=not args.no_audit, chunk_size=chunk_size,
                          timeout_s=args.batch_timeout))
            print(f"   ✓ batch produced {len(precomputed)} unit result(s)")
        else:
            print("\n🧺 BATCH MODE — nothing to batch (all cached/narrative)")

    n_workers = 1 if args.section else args.workers
    print(f"\n⚡ extracting {len(ordered_groups)} group(s) with {n_workers} worker(s)")

    group_results: dict[str, dict] = {}

    with ThreadPoolExecutor(max_workers=n_workers) as pool:
        futures = {
            pool.submit(_extract_group, gnum, gunits, group_leaves, leaf_target): gnum
            for gnum, gunits, group_leaves, leaf_target in ordered_groups
        }
        for fut in as_completed(futures):
            gnum = futures[fut]
            try:
                group_results[gnum] = fut.result()
            except Exception as e:
                print(f"  ❌ group {gnum} failed: {e}")
                group_results[gnum] = {}

    dup_issues = validate_exactly_once(_all_tables_by_unit)
    for iss in dup_issues:
        print(f"   ⚠ {iss}")
    if dup_issues:
        print(f"   ⚠ {len(dup_issues)} duplicate table(s) detected — review before distributing")

    dup_content_issues = flag_duplicate_tables(_all_tables_by_unit)
    for iss in dup_content_issues:
        print(f"   ⚠ {iss}")

    for gnum, gunits, group_leaves, leaf_target in ordered_groups:
        grp_tables = group_results.get(gnum, {})
        for sid, tables in grp_tables.items():
            section_tables[sid].extend(tables)

        print(f"\n##### Section {gnum}  ({len(gunits)} unit(s) -> {len(group_leaves)} subsection tab(s))")

        for lf in group_leaves:
            sid, title = lf["section_id"], lf["title"]
            tables = section_tables.get(sid, [])

            if not tables and _tab_exists(sid):
                sib_unit = next(
                    (u for u in build_units(sections)
                     if any(lf["section_id"] == sid for lf in u["leaves"])),
                    None
                )
                if sib_unit:
                    ext = load_cached_unit(sib_unit, args)
                    if ext is not None:
                        tables = _apply_transforms(ext.tables)
                        print(f"   • {sid} — sibling, reloaded from audit cache")

            if not tables:
                print(f"   · {sid} '{title[:34]}' — no tables, no tab")
                idx = [e for e in idx if e.get("section_id") != sid]
                continue
            if "__FAILED__" in tables:
                print(f"   ❌ {sid} '{title[:34]}' — failed (503/timeout) — re-run to fill")
                continue
            tables = [t for t in tables if t != "__FAILED__"]
            if not tables:
                print(f"   · {sid} '{title[:34]}' — no tables, no tab")
                idx = [e for e in idx if e.get("section_id") != sid]
                continue
            pages_str = (f"{lf['start_page']}" if lf["start_page"] == lf["end_page"]
                         else f"{lf['start_page']}-{lf['end_page']}")
            written_tables: list = []
            for t in tables:
                if not t.columns and all(r.row_type == "note" for r in t.rows):
                    if written_tables:
                        written_tables[-1].rows.extend(t.rows)
                        print(f"   · {sid} — footnote-only table stitched onto '{written_tables[-1].title[:30]}'")
                    continue
                written_tables.append(t)
            total = len(written_tables)
            # A section's tabs are named either '<sid> Table N' (multi-table),
            # '<sid> - <title>' (sheet_name), the bare '<sid>', or — now that a
            # single-table section drops the ' Table N' suffix — the truncated
            # '<sid>'[:31]. Match all forms so a re-run of the same doc doesn't
            # leave a stale tab from the previous naming (e.g. single<->multi).
            _single = sid[:31].rstrip()
            for existing_sname in list(wb.sheetnames):
                if (existing_sname.startswith(f"{sid} Table ") or
                        existing_sname.startswith(f"{sid} -") or
                        existing_sname == sid or existing_sname == _single):
                    wb.remove(wb[existing_sname])
                    used_names.discard(existing_sname)
            for ti, t in enumerate(written_tables, start=1):
                table_label = t.title or f"Table {ti}"
                sname = table_sheet_name(used_names, sid, ti, total)
                ws = wb.create_sheet(title=sname)
                cursor = 4
                cursor = write_table(ws, cursor, t)
                write_section_header(ws, sid, title, ws.max_column,
                                     table_label=table_label, table_n=ti, total_tables=total)
                style_sheet_columns(ws)
                idx = update_index(idx, {"section_id": sid, "title": title, "sheet": sname,
                                         "pages": pages_str, "first_page": int(lf["start_page"]),
                                         "table_n": ti, "n_tables": total})
                print(f"   📄 tab '{sname}'  — {t.title[:40]}")

        stale = []
        for sname in list(wb.sheetnames):
            if sname in ("Contents", "Cost"):
                continue
            if re.search(r'\(\d+\)$', sname):
                stale.append(sname)
            else:
                ws = wb[sname]
                has_data = any(
                    ws.cell(r, c).value
                    for r in range(5, min((ws.max_row or 0) + 1, 20))
                    for c in range(5, (ws.max_column or 0) + 1)
                )
                if not has_data and (ws.max_row or 0) <= 6:
                    stale.append(sname)
        for sname in stale:
            wb.remove(wb[sname])
            used_names.discard(sname)
            print(f"   🗑  removed stale tab '{sname}'")

        save_index(idx)
        _PART_SORT = {None: 0, "A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
        def _sort_key(name):
            if name == "Contents": return (-1, 0, [], 0)
            if name == "Cost":     return (9999, 0, [], 0)
            m = re.match(r'^([A-Z])\.(\d+(?:\.\d+)*)(?:\s+Table\s+(\d+))?', name)
            if m:
                part  = _PART_SORT.get(m.group(1), 9)
                parts = [int(x) for x in m.group(2).split(".")]
                tnum  = int(m.group(3)) if m.group(3) else 0
                return (0, part, parts, tnum)
            m = re.match(r'^(\d+(?:\.\d+)*)(?:\s+Table\s+(\d+))?', name)
            if m:
                parts = [int(x) for x in m.group(1).split(".")]
                tnum  = int(m.group(2)) if m.group(2) else 0
                return (0, 0, parts, tnum)
            return (1, 0, [], 0)
        ordered_names = sorted(wb.sheetnames, key=_sort_key)
        for i, sname in enumerate(ordered_names):
            current_pos = wb.sheetnames.index(sname)
            if current_pos != i:
                wb.move_sheet(sname, offset=i - current_pos)

        rebuild_contents(wb, idx)
        wb.save(out_path)
        print(f"   💾 saved {out_path}  |  run so far: {schema._run_usage['calls']} calls, "
              f"think={schema._run_usage['thinking']:,} tok, ≈ ${schema._run_usage['cost']:.2f}")

        if not args.no_pause and not args.section:
            ans = input("   ⏸  Review the tabs. Enter to continue, 'q' to stop: ").strip().lower()
            if ans == "q":
                print("Stopped by user.")
                break

    if schema._call_log:
        write_cost_sheet(wb, schema._call_log, schema._run_usage, out_path)
        rebuild_contents(wb, idx)
        wb.save(out_path)
        save_cost_summary(schema._call_log, schema._run_usage, out_path, summary_path=str(rp.cost),
                          document=os.path.basename(args.pdf))
        append_to_api_log(schema._call_log, bank or bank_slug.upper(), str(rp.api_log),
                          period=period, document=os.path.basename(args.pdf))

    u = schema._run_usage
    print(f"\n🎉 Done. Workbook: {out_path}")
    print(f"   Total: {u['calls']} calls, input={u['prompt']:,} output={u['output']:,} "
          f"thinking={u['thinking']:,} tok, ≈ ${u['cost']:.4f}")


if __name__ == "__main__":
    main()
